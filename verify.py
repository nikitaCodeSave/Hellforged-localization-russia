#!/usr/bin/env python3
"""Self-contained integrity check for the Hellforged RU localization kit.

Validates the repo's own files (no external dependencies on the analysis project):
  - each csv/<Collection>.csv has the exact header Key,Id,English(en),Russian(ru)
  - every row parses, has a non-empty Russian(ru) and a unique Id
  - SmartFormat argument names ({0}, {amount}) and <tags> in English are preserved in Russian
  - if openpyxl is installed, xlsx sheet row counts match the CSVs

Exit code 0 = all green, 1 = problems (printed). Used as a pre-push hook.

Usage:  python verify.py
"""
from __future__ import annotations
import csv, re, sys, glob, os

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
HEADER = ["Key", "Id", "English(en)", "Russian(ru)"]
ARG = re.compile(r"\{(\w+)")          # {0}, {amount}, {amount:plural:...}
TAG = re.compile(r"<[^<>]+>")          # <b>, <sprite=...>, <1>


def toks(s: str):
    return tuple(sorted(set(ARG.findall(s)))), tuple(sorted(TAG.findall(s)))


def main() -> int:
    problems, total = [], 0
    counts = {}
    csv_files = sorted(glob.glob(os.path.join(HERE, "csv", "*.csv")))
    if not csv_files:
        print("No csv/ files found.")
        return 1

    for path in csv_files:
        coll = os.path.splitext(os.path.basename(path))[0]
        with open(path, encoding="utf-8", newline="") as fh:
            rd = csv.DictReader(fh)
            if rd.fieldnames != HEADER:
                problems.append(f"{coll}: header {rd.fieldnames} != {HEADER}")
                continue
            ids = set()
            n = 0
            for row in rd:
                n += 1
                i = row["Id"]
                if i in ids:
                    problems.append(f"{coll}/{i}: duplicate Id")
                ids.add(i)
                if not row["Russian(ru)"].strip():
                    problems.append(f"{coll}/{i} ({row['Key']}): empty Russian")
                    continue
                if toks(row["English(en)"]) != toks(row["Russian(ru)"]):
                    problems.append(
                        f"{coll}/{i} ({row['Key']}): token mismatch "
                        f"en={toks(row['English(en)'])} ru={toks(row['Russian(ru)'])}")
            counts[coll] = n
            total += n

    # optional xlsx cross-check
    xlsx = os.path.join(HERE, "xlsx", "hellforged-ru.xlsx")
    if os.path.exists(xlsx):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, read_only=True)
            for coll, n in counts.items():
                if coll in wb.sheetnames:
                    rows = wb[coll].max_row - 1  # minus header
                    if rows != n:
                        problems.append(f"xlsx[{coll}]: {rows} rows != csv {n}")
        except ImportError:
            print("(openpyxl not installed — skipping xlsx cross-check)")
        except Exception as e:
            problems.append(f"xlsx: cannot open ({e})")

    print(f"Checked {total} strings across {len(counts)} collections "
          f"({', '.join(f'{k}:{v}' for k, v in counts.items())}).")
    if problems:
        print(f"\nFAILED — {len(problems)} problem(s):")
        for p in problems[:60]:
            print("  -", p)
        return 1
    print("OK — headers valid, ru complete, placeholders/markup preserved, ids unique, xlsx matches.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
